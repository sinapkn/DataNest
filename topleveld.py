import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import jdatetime
import os
from datetime import datetime
from i18n import (
    t, get_language, set_language, Language,
    get_excel_headers, get_date_format
)

filename = "database.xlsx"
top_window = None
selected_item = None


def toplevel_1(root):
    global top_window, selected_item

    if top_window and top_window.winfo_exists():
        top_window.lift()
        return

    top_window = tk.Toplevel(root)
    top_window.title(t("toplevel_title"))
    top_window.geometry("1200x500")
    top_window.resizable(True, True)

    def on_close():
        global top_window, selected_item
        top_window.destroy()
        top_window = None
        selected_item = None

    top_window.protocol("WM_DELETE_WINDOW", on_close)

    top_window.columnconfigure(0, weight=3)
    top_window.columnconfigure(1, weight=2)

    # Left frame - Records
    frame_left = tk.LabelFrame(
        top_window, text=t("frame_records"), font=("Arial", 11, "bold")
    )
    frame_left.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

    # Treeview with scrollbar
    tree_frame = tk.Frame(frame_left)
    tree_frame.pack(expand=True, fill="both")

    tree = ttk.Treeview(
        tree_frame,
        columns=("id", "name", "family", "age", "date"),
        show="headings",
    )
    tree["displaycolumns"] = ("date", "age", "family", "name", "id")

    scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    scrollbar_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

    tree.pack(side="left", expand=True, fill="both")
    scrollbar_y.pack(side="right", fill="y")
    scrollbar_x.pack(side="bottom", fill="x")

    # Set column headings based on language
    headings = [
        ("id", t("col_id")),
        ("name", t("col_name")),
        ("family", t("col_family")),
        ("age", t("col_age")),
        ("date", t("col_date")),
    ]

    for col, txt in headings:
        tree.heading(col, text=txt, anchor=tk.E)
        tree.column(col, anchor=tk.E)
        if col in ("id", "age"):
            tree.column(col, width=50)

    # Right frame - Form
    frame_right = tk.LabelFrame(
        top_window, text=t("frame_form"), font=("Arial", 11, "bold")
    )
    frame_right.grid(row=0, column=1, padx=10, pady=10, sticky="n")

    # Form fields
    tk.Label(frame_right, text=t("label_name")).grid(row=0, column=0, sticky=tk.E, pady=8)
    entry_name = tk.Entry(frame_right, width=25)
    entry_name.grid(row=0, column=1)
    entry_name.bind("<Return>", lambda event: entry_family.focus())

    tk.Label(frame_right, text=t("label_family")).grid(row=1, column=0, sticky=tk.E, pady=8)
    entry_family = tk.Entry(frame_right, width=25)
    entry_family.grid(row=1, column=1)
    entry_family.bind("<Return>", lambda event: entry_age.focus())

    tk.Label(frame_right, text=t("label_age")).grid(row=2, column=0, sticky=tk.E, pady=8)
    entry_age = tk.Entry(frame_right, width=25)
    entry_age.grid(row=2, column=1)
    entry_age.bind("<Return>", lambda event: save_new())

    def load_data():
        tree.delete(*tree.get_children())
        if not os.path.exists(filename):
            return
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                tree.insert("", tk.END, values=row)
            num_rows = len(tree.get_children())
            tree["height"] = min(num_rows, 20)
        except Exception as e:
            messagebox.showerror(t("error_file_operation"), str(e))

    def clear_form():
        entry_name.delete(0, tk.END)
        entry_family.delete(0, tk.END)
        entry_age.delete(0, tk.END)
        entry_name.focus()
        global selected_item
        selected_item = None

    def on_select(event):
        global selected_item
        selected_item = tree.focus()
        values = tree.item(selected_item, "values")
        if values:
            entry_name.delete(0, tk.END)
            entry_family.delete(0, tk.END)
            entry_age.delete(0, tk.END)

            entry_name.insert(0, values[1])
            entry_family.insert(0, values[2])
            entry_age.insert(0, values[3])

    tree.bind("<<TreeviewSelect>>", on_select)
    tree.bind("<Delete>", lambda event: delete_record())

    def get_next_id(ws):
        ids = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
        return max(ids, default=0) + 1

    def validate_input(name, family, age):
        """Validate form input. Returns (is_valid, error_message_key)."""
        if not name or not family:
            return False, "error_invalid_data"
        if not age.isdigit():
            return False, "error_invalid_age"
        age_int = int(age)
        if age_int < 1 or age_int > 150:
            return False, "error_invalid_age_range"
        return True, None

    def get_current_date():
        """Get current date string based on language."""
        if get_language() == Language.PERSIAN:
            return jdatetime.datetime.now().strftime(get_date_format())
        return datetime.now().strftime(get_date_format())

    def save_new():
        name = entry_name.get().strip()
        family = entry_family.get().strip()
        age = entry_age.get().strip()

        is_valid, error_key = validate_input(name, family, age)
        if not is_valid:
            messagebox.showerror(t(error_key), t(error_key))
            return

        date = get_current_date()

        try:
            if os.path.exists(filename):
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(get_excel_headers())

            new_id = get_next_id(ws)
            ws.append([new_id, name, family, age, date])
            wb.save(filename)
        except Exception as e:
            messagebox.showerror(t("error_save_failed"), str(e))
            return

        entry_name.focus()
        load_data()
        clear_form()

    def update_record():
        global selected_item
        if not selected_item:
            messagebox.showwarning(
                t("warning_no_selection"), t("warning_no_selection")
            )
            return

        name = entry_name.get().strip()
        family = entry_family.get().strip()
        age = entry_age.get().strip()

        is_valid, error_key = validate_input(name, family, age)
        if not is_valid:
            messagebox.showerror(t(error_key), t(error_key))
            return

        old_values = tree.item(selected_item, "values")
        if old_values[1] == name and old_values[2] == family and old_values[3] == age:
            messagebox.showinfo(t("info_no_changes"), t("info_no_changes"))
            return

        if not messagebox.askyesno(t("confirm_update"), t("confirm_update")):
            return

        item_id = int(old_values[0])
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if row[0].value == item_id:
                    row[1].value = name
                    row[2].value = family
                    row[3].value = age
                    break

            wb.save(filename)
        except Exception as e:
            messagebox.showerror(t("error_update_failed"), str(e))
            return

        load_data()
        clear_form()

    def delete_record():
        global selected_item
        if not selected_item:
            messagebox.showwarning(
                t("warning_no_selection"), t("warning_no_selection")
            )
            return

        if not messagebox.askyesno(t("confirm_delete"), t("confirm_delete")):
            return

        item_id = int(tree.item(selected_item, "values")[0])
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == item_id:
                    ws.delete_rows(i)
                    break

            wb.save(filename)
        except Exception as e:
            messagebox.showerror(t("error_delete_failed"), str(e))
            return

        load_data()
        clear_form()

    # Buttons frame
    frame_btn = tk.Frame(frame_right)
    frame_btn.grid(row=3, column=0, columnspan=2, pady=15, sticky="ew")
    frame_btn.columnconfigure(0, weight=1)
    frame_btn.columnconfigure(1, weight=1)
    frame_btn.columnconfigure(2, weight=1)
    frame_btn.columnconfigure(3, weight=1)

    btn_width = 12
    tk.Button(
        frame_btn, text=t("btn_save_new"), bg="#51FF00", width=btn_width, command=save_new
    ).grid(row=0, column=0, padx=3, sticky="ew")
    tk.Button(
        frame_btn, text=t("btn_update"), bg="#03D2EE", width=btn_width, command=update_record
    ).grid(row=0, column=1, padx=3, sticky="ew")
    tk.Button(
        frame_btn, text=t("btn_delete"), bg="#FF051E", width=btn_width, command=delete_record
    ).grid(row=0, column=2, padx=3, sticky="ew")
    tk.Button(
        frame_btn, text=t("btn_clear"), bg="#CCCCCC", width=btn_width, command=clear_form
    ).grid(row=0, column=3, padx=3, sticky="ew")

    load_data()